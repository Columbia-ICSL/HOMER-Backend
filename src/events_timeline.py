#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 21 19:11:16 2019

@author: hugomeyer
"""



import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from string import ascii_uppercase
from start_cut import Start_cut
from end_cut import End_cut
from copy import deepcopy
from moviepy.video.io.VideoFileClip import VideoFileClip
import os


class Events_timeline(object):
    def __init__(self, signals, features, duration, fps=10, ground_truth=None):
        self.signals = signals
        self.features = features
        
        self.duration = duration
        self.fps = fps
        self.time = np.linspace(0, self.duration, round(self.duration*self.fps))
        self.ground_truth = ground_truth
        self.sound_dict = {0: 'laugh', 1: 'speech', 2: 'misc'}
        
        self.time_features = deepcopy(self.features)
        self.best_hls = None
        self.many_bests_hls = None
        
        print(self.features['happiness'])
        print()
        self.build_event_timeline()
        print(self.features['happiness'])
        
        
        
        
    def build_event_timeline(self):

        self.signals = dict((key, self.resize_to_fixed_fps(signal, key)) for key, signal in self.signals.items()) 
   
        
        self.reconstruct_emotion_simplified_signal('surprise')
        self.reconstruct_emotion_simplified_signal('happiness')
        
        
        self.reajust_similarity_features()
        self.reajust_sound_features()
        
        self.from_time_to_index()
        self.emotion_preprocessing()
        
        
        
        
        
        
    def compute_highlight(self, hl_min_size=-1, hl_max_size=-1, many_hls=False, hl_overlap_ratio=0.25, 
                         hl_duration_ratio=0.33, max_hl_nb=3, rm_tresh_score=0.15):
        

        hl_min_size = self.t_to_i(hl_min_size) if hl_min_size!=-1 else hl_min_size
        hl_max_size = self.t_to_i(hl_max_size) if hl_max_size!=-1 else hl_max_size
           
        best_hl = self.find_best_hl(hl_max_size, hl_min_size)[0]
        self.best_hls = [best_hl]

        if many_hls:
            best_hls = []
            score = 10
            next_hl = best_hl
            cumul_duration = best_hl['end']-best_hl['start']
            
            
           
            while not best_hls or not (poor_score or score_stuck or cumul_duration_too_long or too_many_hls):
            #(not best_hls or (best_hls[0]['score'] < 1.5 or score > 0.8) or (best_hls[0]['score']  > 1.5 or score > 0.4)) and score != next_hl['score']:
                best_hls.append(next_hl)
                last_score = next_hl['score']
       
                
                for emotion in ['happiness', 'surprise']:
                    to_rm = []
                    for i, feat in enumerate(next_hl['start_cut'].features[emotion]):
                        if feat['rise_start'] < next_hl['start_cut'].end_cut:
                            emo_start = self.time[feat['emo_mask_start']]
                            emo_end = self.time[feat['emo_mask_end']-1]
                            if emo_start < next_hl['end'] and emo_end > next_hl['end']:
                                to_rm.append(i)
                    self.features[emotion] = np.delete(self.features[emotion], to_rm).tolist()
                
                hls = self.find_best_hl(hl_max_size, hl_min_size)
                overlap=True
                i = 0
                while overlap and i < len(hls):
                    overlap = False
                    next_hl = hls[i]
                    for prev_hl in best_hls:
                        both_hls = [prev_hl, next_hl]
                        starts = [prev_hl['start'], next_hl['start']]
                        ends = [prev_hl['end'], next_hl['end']]
                        durations = [prev_hl['end']-prev_hl['start'], next_hl['end']-next_hl['start']]
                        diff = both_hls[np.argmin(starts)]['end'] - both_hls[1-np.argmin(starts)]['start']

                        if diff > 0.25 * min(durations):
                            overlap = True
                    i += 1                
                cumul_duration += next_hl['end']-next_hl['start']
                #print('cumul duration: {:.2f}| duration: {:.2f}' .format(cumul_duration, self.duration))
                poor_score = next_hl['score'] < best_hls[0]['score']/3
                #(best_hls[0]['score']/3 > 1.5 and next_hl['score'] < 0.8) or next_hl['score'] < 0.4
                score_stuck = last_score == next_hl['score'] 
                cumul_duration_too_long = cumul_duration > self.duration*hl_duration_ratio
                too_many_hls = len(best_hls)+1 > max_hl_nb
        else:
            best_hls = [best_hl]
            #best_hl['end_cut'].plot_score_fct(best_hl['end'])
            #best_hl['start_cut'].plot_score_fct()
            
        self.best_hls = best_hls if best_hls[0]['score'] >= rm_tresh_score else []

        if self.best_hls:
            for hl in self.best_hls:
                print('BEST HL: start: {:.2f} | end: {:.2f} | score: {:.2f}'
                      .format(hl['start'], hl['end'], hl['score']))
        else:
            print('NO RELEVANT HIGHLIGHT IN THE VIDEO')
                
        
        
    def emotion_preprocessing(self, two_consec_peak_diff=0.25, surp_hap_diff=1):
        #Give more importance to the surprise emotion than happiness
        #Remove happiness peak if right after a surprise peak
        if self.features['surprise']:
            self.features['surprise'] = [dict((k, v if k!='score' else 1) for k, v in feat.items())
                                        for feat in self.features['surprise']]
            self.features['happiness'] = [dict((k, v if k!='score' else 0 
                                                if (feat2['fall_end'] >= feat1['rise_start']
                                                and feat2['fall_end'] <= feat1['fall_end'])
                                                or abs(feat2['fall_end'] - feat1['rise_start'])<surp_hap_diff
                                                else v/2) 
                                                for k, v in feat1.items()
                                                for feat2 in self.features['surprise'])
                                         for feat1 in self.features['happiness']]
            
        #Merge consecutive peaks, with max of 0.25 seconds time distance
        for emotion in ['happiness', 'surprise']:
            peaks = self.features[emotion]
            to_rm = []
            '''
            for i in range(len(self.features[emotion])-1):
                peak1_end = self.features[emotion][i]['fall_end']
                peak2_start = self.features[emotion][i+1]['rise_start']
   
                if abs(peak1_end - peak2_start) < two_consec_peak_diff*self.fps:
                    self.features[emotion][i]['end'] = self.features[emotion][i+1]['end']
                    self.features[emotion][i]['fall_end'] = self.features[emotion][i+1]['fall_end']
                    w1 = self.features[emotion][i]['fall_end'] - self.features[emotion][i]['rise_start']
                    w2 = self.features[emotion][i+1]['fall_end'] - self.features[emotion][i+1]['rise_start']
                    score1 = self.features[emotion][i]['score']
                    score2 = self.features[emotion][i+1]['score']
                    self.features[emotion][i]['score'] = (w1*score1 + w2*score2)/(w1 + w2)
                    to_rm.append(i+1)
            '''
            i = 0
            while i < len(peaks)-1:
                peak1_end = peaks[i]['fall_end']
                peak2_start = peaks[i+1]['rise_start']
                if abs(peak1_end - peak2_start) < two_consec_peak_diff*self.fps:
                    peaks[i]['end'] = peaks[i+1]['end']
                    peaks[i]['fall_end'] = peaks[i+1]['fall_end']
                    w1 = peaks[i]['fall_end'] - peaks[i]['rise_start']
                    w2 = peaks[i+1]['fall_end'] - peaks[i+1]['rise_start']
                    score1 = peaks[i]['score']
                    score2 = peaks[i+1]['score']
                    peaks[i]['score'] = (w1*score1 + w2*score2)/(w1 + w2)
                    del peaks[i+1]
                    i-=1
                i+=1
            
                
            self.features[emotion] = peaks

        
    def pair_similarity_scores(self):
        labels = [feat['label'] for feat in self.features['similarity']]
        
        for i in range(len(labels)):
            if labels[i] == 'plateau_start':
                j = i+1
                while j<len(labels) and labels[j] != 'plateau_end':
                    j+=1
                if j != len(labels):
                    score1 = self.features['similarity'][i]['score']
                    score2 = self.features['similarity'][j]['score']
                    self.features['similarity'][i]['score'] = max(score1, score2)#(score1+score2)/2
                    self.features['similarity'][j]['score'] = max(score1, score2)#(score1+score2)/2
                    
            if labels[i] == 'valley_start':
                j = i+1
                while j<len(labels) and labels[j] != 'valley_end':
                    j+=1
                if j != len(labels):
                    score1 = self.features['similarity'][i]['score']
                    score2 = self.features['similarity'][j]['score']
                    self.features['similarity'][i]['score'] = max(score1, score2)
                    self.features['similarity'][j]['score'] = max(score1, score2)
                    
                    
    
    def merge_speech(self):
        if self.features['sound']:
            i = 0
            to_rm = []
            nb_sound_events = len(self.features['sound'])
            while i < nb_sound_events-1:
                curr_label_is_speech = self.features['sound'][i]['label'] == 1
                next_label_is_speech = self.features['sound'][i+1]['label'] == 1
                both_events_are_close = (self.features['sound'][i+1]['start']-self.features['sound'][i]['end'])<0.5
                merge_is_not_too_long = (self.features['sound'][i+1]['end']-self.features['sound'][i]['start'])<8
                if curr_label_is_speech and next_label_is_speech and both_events_are_close and merge_is_not_too_long:
                    t1 = self.features['sound'][i]['end']-self.features['sound'][i]['start']
                    t2 = self.features['sound'][i+1]['end']-self.features['sound'][i+1]['start']
                    score1 = self.features['sound'][i]['score']
                    score2 = self.features['sound'][i+1]['score']
                    self.features['sound'][i]['end'] = self.features['sound'][i+1]['end']
                    self.features['sound'][i]['score'] = (t1*score1+t2*score2)/(t1+t2)
                    del self.features['sound'][i+1]
                    i -= 1
                    nb_sound_events -= 1
                i += 1
                    
    
            
            
    def find_best_hl(self, hl_max_size, hl_min_size):
        best_hls = []
        end_cut = End_cut(self.features, self.time, self.fps, self.sound_dict,
                              hl_max_size, hl_min_size)
        end_cut.find_n_best_end_cut_intervals()
        print(end_cut.best_intervals)
        #end_cut.plot_score_fct()
        for i, interval in enumerate(end_cut.best_intervals):
            hls = []
            #print(interval)
            for cut in interval:
                start_cut = Start_cut(self.features, self.time, self.fps, self.sound_dict, 
                                   hl_max_size, hl_min_size)
                start_cut.find_best_start_cut(cut)
                hl = dict()
                hl['start'] = start_cut.best['time']
                hl['end'] = cut['time']
                hl['start_i'] = start_cut.best['index']
                hl['end_i'] = cut['index']
                hl['score'] = (cut['score']+start_cut.best['score'])/2
                hl['score_end'] = cut['score']
                hl['score_start'] = start_cut.best['score']
                hl['end_score_fct'] = end_cut.score_fct
                hl['start_score_fct'] = start_cut.score_fct
                hl['start_cut'] = start_cut
                hl['end_cut'] = end_cut
                hls.append(hl)
                
                #print(hl['score'])
                #end_cut.plot_score_fct(cut['time'])
                #start_cut.plot_score_fct()

            if hls:
                interval_min_score = min([hl['score_start'] for hl in hls])
                interval_max_score = max([hl['score_start'] for hl in hls])
                
                if end_cut.intervals_on_similarity_zone[i]:
                    best_hl_in_interval = hls[int(len(hls)/2)]
                elif interval_max_score-interval_min_score < 0.1:
                    best_hl_in_interval = hls[0]
                else:
                    best_hl_in_interval = hls[np.argmax([hl['score_start'] for hl in hls])]
                #print(best_hl_in_interval)
                best_hls.append(best_hl_in_interval)
        return np.asarray(best_hls)[np.argsort([hl['score'] for hl in best_hls])[::-1]]#[np.argmax([hl['score'] for hl in best_hls])]
            
    def plot_hl(self, save=False):
        if not self.best_hls:
            return 0
        title = 'Score functions for the end and start cuts of the highlight'
        x_label = 'Time'
        y_label = 'Score'
            
        x = self.time#+self.offset
        y = self.best_hls[0]['end_score_fct']
        y2 = self.best_hls[0]['start_score_fct']
    
        fig, ax = plt.subplots(figsize=(10, 4), dpi=80)
        
        factor = 1/self.fps
        
        if x.shape[0] > 400:
            step = round(50*factor)
        elif x.shape[0] > 200:
            step = round(20*factor)
        elif x.shape[0] > 100:
            step = round(10*factor)
        elif x.shape[0] > 30:
            step = max(round(5*factor, 1), 0.1)
        else:
            step = max(round(1*factor, 1), 0.1)
        
        X_ticks_maj = np.arange(max(int(x[0])-factor, 0), int(x[-1])+2, step)
        
        min_ = round(min(min([hl['start_score_fct'][hl['start_i']:hl['end_i']+1].min() 
                          for hl in self.best_hls]), y.min()), 1)-0.2
        max_ = round(max(y.max(), max([hl['start_score_fct'].max() for hl in self.best_hls])), 1)+0.2
        
        y_ticks = np.arange(min_, max_, 0.2)
        X_ticks_min = np.arange(x[0]-factor, x[-1]+1, step*0.2)
        X_ticks_min[0] = int(x[0])

        
        ax.spines['top'].set_color('white')
        ax.spines['right'].set_color('white')
        
        ax.set_xticks(X_ticks_maj)
        ax.set_xticks(X_ticks_min, minor=True)
        ax.set_yticks(y_ticks)
        ax.set_ylim(min_, max_)
        ax.grid(which='both')
        ax.grid(which='minor', alpha=0.4)
        ax.grid(which='major', alpha=1)
        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label)
        plt.title(title)
            
        plt.plot(x, y)
        for hl in self.best_hls:
            plt.plot(x, hl['start_score_fct'])
        
        min_ = min(y.min(), y2.min())
        max_ = max(y.max(), y2.max())
        
        for i, hl in enumerate(self.best_hls):
            label = 'Highlight' if i == 0 else None
            ax.axvline(x=hl['start'], linewidth=1, color='g')
            ax.axvline(x=hl['end'], linewidth=1, color='g')
            ax.fill_between([hl['start'], hl['end']], [min_, min_], [max_, max_], 
                            facecolor='green', alpha=0.1, label=label)

            
        if save == True:
            plt.savefig('Clip_plot_hl.png')
        else:
            plt.show()
            
            
            
            
    def plot(self, name='', save=False, export_path='../'):
        fig, axes = plt.subplots(2, figsize=(10 , 8), dpi=80, sharex=False)
        
        
        
        for i, ax in enumerate(axes):
            
            ax.set_xlabel('Time (s)')
            ax.set_ylabel('Amplitude')
        
            X_ticks_maj = np.arange(0, self.time.shape[-1]+1, 1)
            X_ticks_min = np.arange(0, self.time.shape[-1]+0.2, 0.2)
            y_ticks = np.linspace(0, 1, 11)
            ax.set_xticks(X_ticks_maj)
            ax.set_yticks(y_ticks)
            ax.set_ylim(0, 1)
        
            ax.set_xticks(X_ticks_min, minor=True)
            ax.set_xticks(X_ticks_maj, minor=False)
          
            ax.grid(which='both')
            ax.grid(which='major', alpha=1)
            ax.grid(which='minor', alpha=0.4)
            #for start_end in labels:
             #   plt.axvline(x=start_end[0], linewidth=1, color='r')
             #   plt.axvline(x=start_end[1], linewidth=1, color='r')
             #   mini, maxi = 0, 1
             #   plt.fill_between([start_end[0], start_end[1]], [mini, mini], [maxi, maxi], facecolor='red', alpha=0.05) $
             
            
            if i == 0:
                ax.set_title('Event timeline feature signals along the video')
                
                signals = [self.signals['happiness'], self.signals['surprise']]
                emo_features = [self.time_features['happiness'], self.time_features['surprise']]
                labels = ['Happiness', 'Surprise']
                
                for j, features in enumerate(emo_features):
                    for feat in features:
                        ax.scatter(feat['start'], signals[j][self.t_to_i(feat['start'])], color='b', s=50)
                        ax.scatter(feat['end'], signals[j][self.t_to_i(feat['end'])], color='b', s=50)
                        if feat['rise_start'] is not None:
                            ax.scatter(feat['rise_start'], signals[j][self.t_to_i(feat['rise_start'])], color='r', s=50)
                        if feat['fall_end'] is not None:
                            ax.scatter(feat['fall_end'], signals[j][self.t_to_i(feat['fall_end'])], color='r', s=50)
                
                if self.ground_truth is not None:
                    for (start, end) in self.ground_truth:
                        ax.axvline(x=start, linewidth=1, color='r')
                        ax.axvline(x=end, linewidth=1, color='r')
                        ax.fill_between([start, end], [0, 0], [1, 1], facecolor='red', alpha=0.1)
                
            else:
                for feat in self.time_features['similarity']:
    
                    if any(elem in feat['label'] for elem in ['hill_start', 'hill_end']):
                        color='k'
                    elif any(elem in feat['label'] for elem in ['plateau_start', 'plateau_end']):
                        color='b'
                    elif any(elem in feat['label'] for elem in ['valley_start', 'valley_end']):
                        color='m'
                    else:
                        color='g'
                        
                    size = 20 + feat['score']*150
                    ax.scatter(feat['time'], self.signals['similarity'][self.t_to_i(feat['time'])], color=color, s=size, alpha=0.3)
                    
                speech_flag = False
                laugh_flag = False
                others_flag = False
                    
                for feat in self.time_features['sound']:
                    ax.scatter(feat['start'], self.signals['sound'][self.t_to_i(feat['start'])], color='r', s=50)
                    ax.scatter(feat['end'], self.signals['sound'][self.t_to_i(feat['end'])], color='r', s=50)
                    
                    
                    if feat['label'] == 0:
                        ax.axvline(x=feat['start'], linewidth=1, color='b')
                        ax.axvline(x=feat['end'], linewidth=1, color='b')
                        if not laugh_flag:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='blue', alpha=0.1, label='Laugh') 
                        else:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='blue', alpha=0.1) 
                        laugh_flag = True
                    elif feat['label'] == 1:
                        ax.axvline(x=feat['start'], linewidth=1, color='g')
                        ax.axvline(x=feat['end'], linewidth=1, color='g')
                        if not speech_flag:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='green', alpha=0.1, label='Speech')
                        else:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='green', alpha=0.1)
                        speech_flag = True
                    else:
                        ax.axvline(x=feat['start'], linewidth=1, color='k')
                        ax.axvline(x=feat['end'], linewidth=1, color='k')
                        if not others_flag:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='black', alpha=0.1, label='Miscellaneous')
                        else:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='black', alpha=0.1)
                        others_flag = True

          
                signals = [self.signals['sound'], self.signals['similarity']]
                labels = ['Sound', 'Similarity']

            for signal, label in zip(signals, labels):
                ax.plot(self.time, signal, label=label)
            ax.legend()
        
        if save:
            plt.savefig(export_path+'/event_timeline_{}.png'.format(name))
        else:
            plt.show()
            
            
        
        
    def from_time_to_index(self):
        for emotion in ['happiness', 'surprise']:
            for i in range(len(self.features[emotion])):
                for key in self.features[emotion][i].keys():
                    if key == 'rise_start' and self.features[emotion][i][key] is None:
                        self.features[emotion][i][key] = self.features[emotion][i]['start']
                    if key == 'fall_end' and self.features[emotion][i][key] is None:
                        self.features[emotion][i][key] = self.features[emotion][i]['end']
        for emotion in ['happiness', 'surprise']:
            for i in range(len(self.features[emotion])):
                for key in self.features[emotion][i].keys():
                    if key != 'score':
                        self.features[emotion][i][key] = self.t_to_i(self.features[emotion][i][key])
                  
                    
        for i in range(len(self.features['sound'])):
            self.features['sound'][i]['start'] = self.t_to_i(self.features['sound'][i]['start'])
            self.features['sound'][i]['end'] = self.t_to_i(self.features['sound'][i]['end'])
            
        for i in range(len(self.features['similarity'])):
            self.features['similarity'][i]['time'] = self.t_to_i(self.features['similarity'][i]['time'])
            
            
    def t_to_i(self, ts):
        return np.abs(self.time-ts).argmin()
         
    
    def info(self):
        for name, features in self.features.items():
            print(name)
            for feature in features:
                print(feature)
                
                
            print()

            
            
            
            

            
            
             
    def reconstruct_emotion_simplified_signal(self, label):
        signal = np.zeros(round(self.duration*self.fps))
        for feat in self.features[label]:
            start = self.t_to_i(feat['start'])
            end = self.t_to_i(feat['end'])
            if feat['rise_start'] is not None:
                rise_start = self.t_to_i(feat['rise_start'])
                rise_vals = np.linspace(0, feat['score'], start-rise_start+1)[:-1]
                signal[rise_start:start] = rise_vals
            if feat['fall_end'] is not None:
                fall_end = self.t_to_i(feat['fall_end'])
                fall_vals = np.linspace(feat['score'], 0, fall_end-end+1)[1:]
                signal[end+1:fall_end+1] = fall_vals
            #print(rise_start, fall_end, start, end)
            signal[start:end+1] = feat['score']
            
        self.signals[label] = signal
        
        
    def reajust_similarity_features(self):
        signal = self.signals['similarity']
        for i in range(len(self.features['similarity'])):
            feat = self.features['similarity'][i]
            moving_pts = [self.t_to_i(feat['time']), self.t_to_i(feat['time'])]
            distances = [0, 0]
                
            if moving_pts[0]<1 or moving_pts[0]>=len(signal)-1 or (signal[moving_pts[0]-1]-signal[moving_pts[0]])*(signal[moving_pts[0]+1]-signal[moving_pts[0]])<0:
                
                if moving_pts[0] < 1:
                    distances[0] = float('inf')
                else:                    
                    init_diff = signal[moving_pts[0]-1]-signal[moving_pts[0]]
                    while moving_pts[0] >= 1 and init_diff*(signal[moving_pts[0]-1]-signal[moving_pts[0]])>0:
                        moving_pts[0] -= 1
                        distances[0] += 1
                    if moving_pts[0] == 0:
                        distances[0] = float('inf')
                        
                if moving_pts[1] >= len(signal)-1:
                    distances[1] = float('inf')
                    
                else:                    
                    init_diff = signal[moving_pts[1]+1]-signal[moving_pts[1]]
                    while moving_pts[1]+1 < len(signal) and init_diff*(signal[moving_pts[1]+1]-signal[moving_pts[1]])>0:
                        moving_pts[1] += 1
                        distances[1] += 1
                    if moving_pts[1] == len(signal)-1:
                        distances[1] = float('inf')
                    
                self.features['similarity'][i]['time'] = self.time[moving_pts[np.argmin(distances)]]
                    
        self.pair_similarity_scores()
                
                
            
    def reajust_sound_features(self):
        signal = self.signals['sound']
        for i in range(len(self.features['sound'])):
            feature = self.features['sound'][i]
            feat_start = self.t_to_i(feature['start'])
            feat_end = self.t_to_i(feature['end'])
            
            prev = int((feat_start+feat_end)/2)
            next_ = prev-1
            while prev > 0 and signal[next_] == signal[prev]:
                prev = next_
                next_ -= 1
            self.features['sound'][i]['start'] = self.time[prev]
                
            prev = int((feat_start+feat_end)/2)
            next_ = prev+1
            while next_ < len(signal) and signal[next_] == signal[prev]:
                prev = next_
                next_ += 1
            self.features['sound'][i]['end'] = self.time[prev]
            
        self.merge_speech()
            
            
        
    def resize_to_fixed_fps(self, signal, label):
        #print(len(signal), duration, len(signal)/duration)        
        nb_pts = abs(len(signal)-round(self.duration*self.fps))
        if nb_pts:
            step = int(len(signal)/(nb_pts+1))
            indices = np.linspace(step, len(signal)-step-1, nb_pts).astype(int)
            diff = indices[:-1]-indices[1:]
            if diff[diff == 0].shape[0] != 0:
                #raise ValueError('The signal has an too low fps for a {} fps interpolation.' .format(self.fps))
                pass
            if len(signal)/self.duration > self.fps:
                signal = np.delete(signal, indices).tolist()
            else:
                #print(signal, indices)
                if label == 'sound':
                    values = [signal[i-1] if i else signal[0] for i in indices]
                else:
                    values = [(signal[i] + signal[i-1])/2 if i else signal[0] for i in indices]
                signal = np.insert(signal, indices, values).tolist()
        return signal
    
    
    
    def export(self, ID, path):
        signals = np.asarray([signal for signal in self.signals.values()])
        df = pd.DataFrame(signals.T, columns = self.signals.keys())
        df.insert(0, 'time', self.time)
        
        try: 
            book = load_workbook(path + '/ET_signals.xlsx')
            writer = pd.ExcelWriter(path + '/ET_signals.xlsx', engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, "Clip_{}".format(ID))
            writer.save()
        except:
            df.to_excel(path + '/ET_signals.xlsx', sheet_name="Clip_{}".format(ID))
        
        wb = openpyxl.load_workbook(path + '/ET_signals.xlsx')

        sheet = wb.get_sheet_by_name("Clip_{}".format(ID))
        sheet['G1'] = 'duration'
        sheet['G2'] = self.duration
        wb.save(path + '/ET_signals.xlsx')
        
        
        try: 
            wb = openpyxl.load_workbook(path + '/ET_features.xlsx')
        except:
            wb = openpyxl.Workbook()
            
        try:
            sheet = wb["Clip_{}".format(ID)]
        except: 
            sheet = wb.create_sheet("Clip_{}".format(ID))
        
        
        col = 0
        for name, features in self.features.items():
            row = 1          
            sheet[ascii_uppercase[col]+str(row)] = str(name)            
            row += 1
            
            for feature in features:
                for key, val in feature.items():
                    row += 1
                    sheet[ascii_uppercase[col]+str(row)] = key
                    sheet[ascii_uppercase[col+1]+str(row)] = val
                row += 1
            col += 3
            
        wb.save(path + '/ET_features.xlsx')
                        
    
   
            
    def export_hl(self, input_path, output_path):      
        for i, hl in enumerate(self.best_hls):
            #Correct the error of ffmpeg
            start_time = max(hl['start']-0.1, 0)
            end_time = hl['end']
            hl_path = '.'.join(output_path.split('.')[:-1]) + '_{}'.format(i+1) + '.mp4' if len(self.best_hls) > 1 else output_path
            
            print(hl_path)

            #ffmpeg_extract_subclip(input_path, start_time, end_time, targetname=output_path)
            with VideoFileClip(input_path) as video:
                new = video.subclip(start_time, min(end_time, video.duration))
                new.write_videofile(hl_path, audio_codec='aac')


            path = '/'.join(hl_path.split('/')[:-1])
            start_end_times = pd.DataFrame(np.array([[start_time, end_time]]), columns = ['start_time', 'end_time'])
            start_end_times.to_csv(os.path.join(path, 'start_end_cuts.csv'))
            
            
            
            
    def plot_TL_and_HL(self, name='', save=False, export_path='../'):
        fig, axes = plt.subplots(4, figsize=(10 , 18), dpi=80, sharex=False)
        
        fig.subplots_adjust(hspace=0.2)
        
        for i, ax in enumerate(axes):
            
            
            ax.set_ylabel('Amplitude')
        
            X_ticks_maj = np.arange(0, self.time.shape[-1]+1, 1)
            X_ticks_min = np.arange(0, self.time.shape[-1]+0.2, 0.2)
            
            ax.set_xticks(X_ticks_maj)
        
            ax.set_xticks(X_ticks_min, minor=True)
            ax.set_xticks(X_ticks_maj, minor=False)
          
            ax.grid(which='both')
            ax.grid(which='major', alpha=1)
            ax.grid(which='minor', alpha=0.4)
            #for start_end in labels:
             #   plt.axvline(x=start_end[0], linewidth=1, color='r')
             #   plt.axvline(x=start_end[1], linewidth=1, color='r')
             #   mini, maxi = 0, 1
             #   plt.fill_between([start_end[0], start_end[1]], [mini, mini], [maxi, maxi], facecolor='red', alpha=0.05)  
            if i == 0:
                ax.set_title('Event timeline feature signals along the video')
                
                y_ticks = np.linspace(0, 1, 11)
                ax.set_yticks(y_ticks)
            
                signals = [self.signals['sound'], self.signals['similarity'], self.signals['happiness'], self.signals['surprise']]
                labels = ['Sound', 'Similarity', 'Happiness', 'Surprise']
                
              
                ax.set_ylim(0, 1)
            
                for feat in self.time_features['similarity']:
    
                    if any(elem in feat['label'] for elem in ['hill_start', 'hill_end']):
                        color='k'
                    elif any(elem in feat['label'] for elem in ['plateau_start', 'plateau_end']):
                        color='b'
                    elif any(elem in feat['label'] for elem in ['valley_start', 'valley_end']):
                        color='m'
                    else:
                        color='g'
                        
                    size = 20 + feat['score']*150
                    ax.scatter(feat['time'], self.signals['similarity'][self.t_to_i(feat['time'])], color=color, s=size)
                    
                speech_flag = False
                laugh_flag = False
                others_flag = False
                    
                for feat in self.time_features['sound']:
                    #ax.scatter(feat['start'], self.signals['sound'][self.t_to_i(feat['start'])], color='r', s=50)
                    #ax.scatter(feat['end'], self.signals['sound'][self.t_to_i(feat['end'])], color='r', s=50)
                    
                    
                    if feat['label'] == 0:
                        ax.axvline(x=feat['start'], linewidth=1, color='b')
                        ax.axvline(x=feat['end'], linewidth=1, color='b')
                        if not laugh_flag:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='blue', alpha=0.1, label='Laugh') 
                        else:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='blue', alpha=0.1) 
                        laugh_flag = True
                    elif feat['label'] == 1:
                        ax.axvline(x=feat['start'], linewidth=1, color='g')
                        ax.axvline(x=feat['end'], linewidth=1, color='g')
                        if not speech_flag:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='green', alpha=0.1, label='Speech')
                        else:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='green', alpha=0.1)
                        speech_flag = True
                    else:
                        ax.axvline(x=feat['start'], linewidth=1, color='k')
                        ax.axvline(x=feat['end'], linewidth=1, color='k')
                        if not others_flag:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='black', alpha=0.1, label='Miscellaneous')
                        else:
                            ax.fill_between([feat['start'], feat['end']], [0, 0], [1, 1], facecolor='black', alpha=0.1)
                        others_flag = True


                for signal, label in zip(signals, labels):
                    ax.plot(self.time, signal, label=label)
          
                
            if i == 3:
                ax.set_title('Score functions for the end and start cuts of the highlight')
           
                x = self.time#+self.offset
                y = self.best_hl['end_score_fct']
                y2 = self.best_hl['start_score_fct']
            
               
                
                min_ = min(y.min(), y2.min())
                max_ = max(y.max(), y2.max())
                
                ax.set_ylim(min_, max_)
                
                y_ticks = np.arange(int(min_)-1, int(max_)+2, 0.5)
                ax.set_yticks(y_ticks)
                
                ax.axvline(x=self.best_hl['start'], linewidth=1, color='g')
                ax.axvline(x=self.best_hl['end'], linewidth=1, color='g')
                ax.fill_between([self.best_hl['start'], self.best_hl['end']], [min_, min_], [max_, max_], 
                                facecolor='green', alpha=0.1, label='Highlight')
                
                ax.plot(x, y, label='End cut function')
                ax.plot(x, y2, label='Start cut function')
                
            if i == 1:
                ax.set_title('Score functions associated to the end cut of the highlight')
                x = self.time
                y = np.asarray(self.best_hl['end_cut'].score_fct)
                #ax.axvline(x=self.best_hl['end_cut'].time[np.argmax(y)], linewidth=1, color='g')
                ax.plot(x, y, label='Total')
                ax.plot(x, self.best_hl['end_cut'].emo_score_fct, label='Emotion')
                ax.plot(x, self.best_hl['end_cut'].sound_score_fct, label='Sound')
                ax.plot(x, self.best_hl['end_cut'].simil_score_fct, label='Similarity')
                
            if i == 2:
                ax.set_title('Score functions associated to the start cut of the highlight')
                x = self.time
                y = np.asarray(self.best_hl['start_cut'].score_fct)
                #ax.axvline(x=self.best_hl['end_cut'].time[np.argmax(y)], linewidth=1, color='g')
                ax.plot(x, y, label='Total')
                ax.plot(x, self.best_hl['start_cut'].emo_score_fct, label='Emotion')
                ax.plot(x, self.best_hl['start_cut'].sound_score_fct, label='Sound')
                ax.plot(x, self.best_hl['start_cut'].simil_score_fct, label='Similarity')
                ax.plot(x, self.best_hl['start_cut'].time_score_fct, label='Time')
                ax.set_xlabel('Time (s)')
                
            ax.legend()
        
        if save:
            plt.savefig(export_path)
        else:
            plt.show()

        
        
